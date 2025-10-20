import asyncio
import sys
import os
from typing import Dict, Any, Optional

from openpyxl import load_workbook

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

# Используем объекты приложения
"""Импорт номенклатуры VED из Excel в БД.

Запуск предполагается внутри backend-контейнера, где проект смонтирован в /app.
Чтобы гарантировать доступность модулей проекта (database, models),
добавляем /app в sys.path даже если переменная PYTHONPATH не выставлена.
"""

# Гарантируем доступ к модулям приложения
APP_ROOT = "/app"
if APP_ROOT not in sys.path:
    sys.path.insert(0, APP_ROOT)

from database import async_engine
from models import VEDNomenclature


HEADER_ALIASES: Dict[str, str] = {
    "код": "code_1c",
    "код 1с": "code_1c",
    "code": "code_1c",
    "1c": "code_1c",

    "артикул": "article",
    "article": "article",

    "наименование": "name",
    "наименование товара": "name",
    "name": "name",

    "матрица": "matrix",
    "matrix": "matrix",

    "глубина бурения": "drilling_depth",
    "depth": "drilling_depth",

    "высота": "height",
    "height": "height",

    "резьба": "thread",
    "thread": "thread",

    "тип продукта": "product_type",
    "тип": "product_type",
    "product_type": "product_type",
}


def normalize(s: Optional[str]) -> str:
    return (s or "").strip()


def map_header(name: Any) -> Optional[str]:
    if name is None:
        return None
    key = str(name).strip().lower()
    return HEADER_ALIASES.get(key)


async def upsert_row(db: AsyncSession, row: Dict[str, Any]) -> bool:
    code = normalize(row.get("code_1c"))
    article = normalize(row.get("article"))
    name = normalize(row.get("name"))
    matrix = normalize(row.get("matrix"))
    drilling_depth = normalize(row.get("drilling_depth"))
    height = normalize(row.get("height"))
    thread = normalize(row.get("thread"))
    product_type = normalize(row.get("product_type") or "коронка")

    if not code and not article:
        return False  # пропускаем пустые строки
    if not name or not matrix:
        return False

    # Ищем по code_1c, если есть, иначе по article
    stmt = None
    if code:
        stmt = select(VEDNomenclature).where(VEDNomenclature.code_1c == code)
    else:
        stmt = select(VEDNomenclature).where(VEDNomenclature.article == article)

    result = await db.execute(stmt)
    entity: Optional[VEDNomenclature] = result.scalar_one_or_none()

    if entity is None:
        entity = VEDNomenclature(
            code_1c=code or f"ART-{article}",
            name=name,
            article=article or code,
            matrix=matrix,
            drilling_depth=drilling_depth or None,
            height=height or None,
            thread=thread or None,
            product_type=product_type or "коронка",
            is_active=True,
        )
        db.add(entity)
    else:
        # Обновляем поля
        entity.name = name or entity.name
        entity.article = (article or entity.article)
        entity.matrix = matrix or entity.matrix
        entity.drilling_depth = drilling_depth or entity.drilling_depth
        entity.height = height or entity.height
        entity.thread = thread or entity.thread
        entity.product_type = product_type or entity.product_type

    return True


async def import_file(path: str, sheet: Optional[str] = None) -> None:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    # Читаем заголовки
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [map_header(cell.value) for cell in header_row]
    indices = {i: h for i, h in enumerate(headers) if h}

    # Если ключевых заголовков нет, используем запасное сопоставление по индексам
    must_have = set(["name", "matrix"])  # code/article может быть пустым, но хотя бы name/matrix нужны
    if not indices or not must_have.issubset(set(indices.values())):
        # Фоллбэк сопоставления: A..H
        fallback = {
            0: "code_1c",      # A
            1: "article",      # B
            2: "name",         # C
            3: "matrix",       # D
            4: "drilling_depth",# E
            5: "height",       # F
            6: "thread",       # G
            7: "product_type", # H
        }
        indices = fallback
        print("[IMPORT] Заголовки не распознаны полностью, применён фоллбэк по колонкам A..H")

    print("[IMPORT] Маппинг колонок:", indices)

    total = 0
    created_or_updated = 0

    SessionLocal = async_sessionmaker(async_engine, expire_on_commit=False)

    skipped_empty = 0
    skipped_required = 0

    async with SessionLocal() as db:
        for row in ws.iter_rows(min_row=2):
            total += 1
            data: Dict[str, Any] = {}
            empty = True
            for idx, cell in enumerate(row):
                key = indices.get(idx)
                if not key:
                    continue
                value = cell.value
                if isinstance(value, str):
                    value = value.strip()
                if value not in (None, ""):
                    empty = False
                data[key] = value
            if empty:
                skipped_empty += 1
                continue
            ok = await upsert_row(db, data)
            if ok:
                created_or_updated += 1
            else:
                skipped_required += 1
        await db.commit()

    print(f"Импорт завершен. Обработано строк: {total}, записей создано/обновлено: {created_or_updated}, пропущено пустых: {skipped_empty}, пропущено без обязательных полей: {skipped_required}")


def main():
    if len(sys.argv) < 2:
        print("Использование: python scripts/import_ved_nomenclature.py <path_to_xlsx> [sheet_name]")
        sys.exit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    asyncio.run(import_file(path, sheet))


if __name__ == "__main__":
    main()


